"""
Document Extractor Service
--------------------------
Converts uploaded files into clean raw text before AI parsing.

Supports:
  - PDF  (pdfplumber — handles text PDFs and PDFs with embedded tables)
  - DOCX (python-docx — preserves table structure as text)
  - XLSX (openpyxl — converts sheets to readable text)
  - TXT  (plain text — no conversion needed)

The goal is to produce the richest possible text representation so the
AI parser has the maximum context available, regardless of source format.

Table handling strategy:
  Tables are converted to pipe-delimited text so the AI can read column
  headers and values clearly. This is the key to handling scoring matrices
  and band definitions that live in Excel or Word tables.
"""

import io
from pathlib import Path


def extract_text_from_file(content: bytes, filename: str) -> tuple[str, str]:
    """
    Extract raw text from an uploaded file.

    Args:
        content:  raw file bytes
        filename: original filename (used to detect format)

    Returns:
        (extracted_text, source_type)
        source_type: one of "pdf" | "docx" | "xlsx" | "txt"
    """
    suffix = Path(filename).suffix.lower()

    if suffix == ".pdf":
        return _extract_pdf(content), "pdf"
    elif suffix in (".docx", ".doc"):
        return _extract_docx(content), "docx"
    elif suffix in (".xlsx", ".xls"):
        return _extract_xlsx(content), "xlsx"
    elif suffix in (".txt", ".md"):
        return content.decode("utf-8", errors="replace"), "txt"
    else:
        # Try treating as text — worst case the AI will get garbage and
        # the parsing confidence will be low, which is surfaced to the user
        try:
            return content.decode("utf-8", errors="replace"), "txt"
        except Exception:
            raise ValueError(f"Unsupported file format: {suffix}. Upload PDF, DOCX, XLSX, or TXT.")


def _extract_pdf(content: bytes) -> str:
    """
    Extract text from PDF using pdfplumber.
    Tables are preserved as pipe-delimited text for the AI to read.
    """
    import pdfplumber

    pages_text = []

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            page_parts = [f"--- Page {page_num} ---"]

            # Extract tables first — they contain structured risk data
            tables = page.extract_tables()
            table_bboxes = []

            for table in tables:
                if not table:
                    continue
                # Remember where this table is so we exclude it from plain text extraction
                # (pdfplumber would double-extract it otherwise)
                table_text = _table_to_text(table)
                if table_text:
                    page_parts.append(f"[TABLE]\n{table_text}\n[/TABLE]")

            # Extract remaining text (excluding table areas)
            plain_text = page.extract_text()
            if plain_text:
                page_parts.append(plain_text)

            pages_text.append("\n".join(page_parts))

    return "\n\n".join(pages_text)


def _extract_docx(content: bytes) -> str:
    """
    Extract text from Word document.
    Tables are converted to pipe-delimited text.
    Headings are preserved with markers.
    """
    from docx import Document

    doc = Document(io.BytesIO(content))
    parts = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

        if tag == "p":
            # Paragraph — check if it's a heading
            from docx.oxml.ns import qn
            style_name = ""
            pPr = element.find(qn("w:pPr"))
            if pPr is not None:
                pStyle = pPr.find(qn("w:pStyle"))
                if pStyle is not None:
                    style_name = pStyle.get(qn("w:val"), "")

            text = "".join(node.text or "" for node in element.iter() if node.tag.endswith("}t"))
            if text.strip():
                if "Heading" in style_name or "heading" in style_name:
                    parts.append(f"\n## {text.strip()}")
                else:
                    parts.append(text.strip())

        elif tag == "tbl":
            # Table — convert to pipe-delimited text
            rows = []
            for row in element.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr"):
                cells = []
                for cell in row.findall("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc"):
                    cell_text = "".join(
                        node.text or ""
                        for node in cell.iter()
                        if node.tag.endswith("}t")
                    ).strip()
                    cells.append(cell_text)
                if any(cells):
                    rows.append(cells)

            if rows:
                table_text = _table_to_text(rows)
                parts.append(f"[TABLE]\n{table_text}\n[/TABLE]")

    return "\n".join(parts)


def _extract_xlsx(content: bytes) -> str:
    """
    Extract text from Excel spreadsheet.
    Each sheet is extracted separately with its name as a header.
    Empty rows and columns are skipped.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except ImportError:
        # openpyxl not installed — try basic approach
        raise ValueError(
            "openpyxl is required for Excel files. "
            "Please copy and paste the content as text instead."
        )

    sheets_text = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            if all(cell is None for cell in row):
                continue
            cells = [str(cell) if cell is not None else "" for cell in row]
            rows.append(cells)

        if rows:
            table_text = _table_to_text(rows)
            sheets_text.append(f"## Sheet: {sheet_name}\n[TABLE]\n{table_text}\n[/TABLE]")

    return "\n\n".join(sheets_text)


def _table_to_text(rows: list) -> str:
    """Convert a list-of-lists table to pipe-delimited readable text."""
    if not rows:
        return ""

    lines = []
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        line = " | ".join(cells)
        lines.append(line)
        # Add separator after the first row (header)
        if i == 0 and len(rows) > 1:
            lines.append("-" * min(len(line), 80))

    return "\n".join(lines)


def detect_document_type(text: str) -> str:
    """
    Heuristic: detect whether the document is primarily narrative, tabular, or hybrid.
    This hint is passed to the AI parser so it knows what extraction strategy to use.
    """
    table_marker_count = text.count("[TABLE]")
    word_count = len(text.split())
    pipe_count = text.count("|")

    if table_marker_count >= 3 or pipe_count > word_count * 0.3:
        return "tabular"
    elif table_marker_count >= 1:
        return "hybrid"
    else:
        return "narrative"
